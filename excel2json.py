from openpyxl import load_workbook


class Sheet:
    def __init__(self, wsheet):
        self.wsheet = wsheet

    def __getitem__(self, item):
        cell_value = self.wsheet[item].value
        if cell_value is not None:
            cell_value = str(self.wsheet[item].value).strip()
        return cell_value


def isNoneOrEmpty(entity):
    return (entity is None) or (len(entity) == 0)


def map_headings(ws, heading_row=1, start_scan='A'):
    excel_col_map = {}
    cur_column = start_scan
    head_row = str(heading_row)
    col_ord = ord(cur_column)
    while not isNoneOrEmpty(ws[cur_column + head_row]) and cur_column != 'Z':
        excel_col_map[ws[cur_column + head_row].lower()] = chr(col_ord)
        col_ord += 1
        cur_column = chr(col_ord)
    if cur_column == 'Z':
        print('** Warning: more columns than expected!\n')
    return excel_col_map


def map_row(ws, row_number, sheet_col_map):
    dict_of_row = {}
    for key in sheet_col_map:
        column = sheet_col_map[key]
        cell_value = ws[column + str(row_number)]
        if cell_value and len(cell_value) > 0:
            dict_of_row[key] = cell_value
    return dict_of_row


def map_rows_in_sheet(ws, heading_row=1):
    sheet_col_map = map_headings(ws, heading_row)
    row_number = heading_row + 1
    MAX_ROWS = 10000
    mapped_rows = []
    while row_number < MAX_ROWS:
        dict_of_row = map_row(ws, row_number, sheet_col_map)
        if len(dict_of_row) > 0:
            mapped_rows.append(dict_of_row)
        else:
            break
        row_number += 1
    return mapped_rows


def get_sheet(excel_filename, sheet_number=0):
    w = load_workbook(excel_filename, data_only=True)
    ws = Sheet(w[w.sheetnames[sheet_number]])
    return ws


def rows_to_dict_list(excel_filename, sheet_index=0, heading_row=1):
    ws = get_sheet(excel_filename, sheet_number=sheet_index)
    dict_list = map_rows_in_sheet(ws, heading_row=heading_row)
    if len(dict_list) == 0:
        print(f"Warning: No rows mapped for {excel_filename}")
    return dict_list
